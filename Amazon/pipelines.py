# -*- coding: utf-8 -*-

# Define your item pipelines here
#
# Don't forget to add your pipeline to the ITEM_PIPELINES setting
# See: https://doc.scrapy.org/en/latest/topics/item-pipeline.html
from openpyxl import Workbook
from scrapy.exceptions import DropItem

# class MociPipeline(object):
#     def __init__(self, filename):
#     	self.filename = filename
#     	self.wb = Workbook()
#     	self.sheet = self.wb.active

#     @classmethod
#     def from_crawler(cls, crawler):
#     	filename = crawler.settings.get('MOCI_FILENAME')
#     	return cls(filename)

#     def open_spider(self, spider):
#         title = ['pharse','volume']

#         for field in range(1,len(title)+1):
#             _= self.sheet.cell(row=1,column=field,value=title[field-1])

#     def close_spider(self, spider):
#     	self.wb.save(self.filename)

#     def process_item(self, item, spider):
#         pharse = item['pharse']
#         volume = item['volume']
#         self.sheet.append([pharse,volume])

#         raise DropItem

class AmazonPipeline(object):
    def __init__(self, filename):
    	self.filename = filename
    	self.wb = Workbook()
    	self.sheet = self.wb.active

    @classmethod
    def from_crawler(cls, crawler):
    	filename = crawler.settings.get('AMAZON_FILENAME')
    	return cls(filename)

    def open_spider(self, spider):
        title = ['pharse','volume','mod']

        for field in range(1,len(title)+1):
            _= self.sheet.cell(row=1,column=field,value=title[field-1])

    def close_spider(self, spider):
    	self.wb.save(self.filename)

    def process_item(self, item, spider):
        pharse = item['pharse']
        volume = item['volume']
        mod = item['mod']
        self.sheet.append([pharse,volume,mod])